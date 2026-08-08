"""Versioned Dhan equity tradebook and holdings parsers."""

from __future__ import annotations

import csv
import hashlib
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal, cast
from zoneinfo import ZoneInfo

from openpyxl import load_workbook  # type: ignore[import-untyped]

from ..identity import canonical_json, is_valid_isin
from ..models import ParseResult, ParsedFill, ParsedHolding, RawRow

TRADE_HEADERS = (
    "Symbol", "ISIN", "Trade Date", "Exchange", "Segment", "Series",
    "Trade Type", "Auction", "Quantity", "Price", "Trade ID", "Order ID",
    "Order Execution Time",
)
HOLDING_HEADERS = (
    "Instrument", "Qty.", "Avg. cost", "LTP", "Invested", "Cur. val",
    "P&L", "Net chg.", "Day chg.",
)


def _decimal(value: Any, field: str) -> Decimal:
    if isinstance(value, str) and value.startswith("="):
        raise ValueError(f"formula is not allowed in {field}")
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"invalid decimal for {field}: {value!r}") from exc
    if not result.is_finite():
        raise ValueError(f"non-finite decimal for {field}")
    return result


def _row_hash(values: dict[str, Any]) -> str:
    return hashlib.sha256(canonical_json(values).encode("utf-8")).hexdigest()


def _date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _datetime(value: Any) -> datetime:
    parsed = value if isinstance(value, datetime) else datetime.fromisoformat(str(value))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=ZoneInfo("Asia/Kolkata"))
    return parsed.astimezone(UTC).replace(tzinfo=None)


class DhanTradebookParser:
    format_version = "dhan-equity-tradebook-v1"

    def parse(self, path: Path) -> ParseResult:
        # Dhan's export currently declares an incorrect worksheet dimension
        # (A1:A5276). openpyxl read-only mode trusts it and would hide B:N.
        wb = load_workbook(path, read_only=False, data_only=False)
        try:
            if "Equity" not in wb.sheetnames:
                raise ValueError("Dhan tradebook is missing the Equity sheet")
            sheet = wb["Equity"]
            header_row: int | None = None
            header_values: list[str] = []
            metadata_rows: list[list[Any]] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(v).strip() if v is not None else "" for v in row]
                if all(name in values for name in TRADE_HEADERS):
                    header_row, header_values = row_number, values
                    break
                metadata_rows.append(list(row))
            if header_row is None:
                raise ValueError("Dhan tradebook header signature was not found")
            indexes = {name: header_values.index(name) for name in TRADE_HEADERS}
            raw_rows: list[RawRow] = []
            records: list[ParsedFill] = []
            issues: list[dict[str, Any]] = []
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                source = {
                    name: row[indexes[name]] if indexes[name] < len(row) else None
                    for name in TRADE_HEADERS
                }
                raw_rows.append(RawRow("Equity", row_number, source, _row_hash(source)))
                side = str(source["Trade Type"] or "").strip().lower()
                if side not in {"buy", "sell"}:
                    raise ValueError(f"invalid side at Equity row {row_number}")
                quantity = _decimal(source["Quantity"], "Quantity")
                price = _decimal(source["Price"], "Price")
                if quantity <= 0 or price <= 0:
                    raise ValueError(f"non-positive economics at Equity row {row_number}")
                raw_isin = str(source["ISIN"] or "").strip().upper()
                isin = raw_isin if is_valid_isin(raw_isin) else None
                if isin is None:
                    issues.append({
                        "severity": "ERROR", "issue_type": "MALFORMED_ISIN",
                        "row_number": row_number,
                        "evidence": {"value": raw_isin, "symbol": source["Symbol"]},
                    })
                records.append(ParsedFill(
                    symbol=str(source["Symbol"]).strip().upper(), isin=isin,
                    trade_date=_date(source["Trade Date"]),
                    exchange=str(source["Exchange"]).strip().upper(),
                    segment=str(source["Segment"]).strip().upper(),
                    series=str(source["Series"]).strip().upper(),
                    side=cast("Literal['buy', 'sell']", side),
                    auction=bool(source["Auction"]), quantity=quantity, price=price,
                    trade_id=str(source["Trade ID"]).strip(),
                    order_id=str(source["Order ID"]).strip(),
                    executed_at=_datetime(source["Order Execution Time"]),
                    raw_row_number=row_number,
                ))
            dates = [record.trade_date for record in records]
            metadata_text = " ".join(
                str(value) for row in metadata_rows for value in row if value is not None
            )
            return ParseResult(
                "tradebook", self.format_version, tuple(raw_rows), tuple(records),
                {"sheet": "Equity", "header_row": header_row,
                 "metadata_text": metadata_text,
                 "detected_from": min(dates).isoformat() if dates else None,
                 "detected_to": max(dates).isoformat() if dates else None},
                tuple(issues),
            )
        finally:
            wb.close()


class DhanHoldingsParser:
    format_version = "dhan-holdings-v1"

    def parse(self, path: Path) -> ParseResult:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            matrix = list(csv.reader(handle))
        if not matrix:
            raise ValueError("holdings CSV is empty")
        width = max(len(row) for row in matrix)
        for row in matrix:
            row.extend([""] * (width - len(row)))
        while matrix[0] and all(not str(row[-1]).strip() for row in matrix):
            for row in matrix:
                row.pop()
        headers = [str(value).strip() for value in matrix[0]]
        if not all(name in headers for name in HOLDING_HEADERS):
            raise ValueError(f"unsupported Dhan holdings schema: {headers}")
        indexes = {name: headers.index(name) for name in HOLDING_HEADERS}
        raw_rows: list[RawRow] = []
        records: list[ParsedHolding] = []
        issues: list[dict[str, Any]] = []
        for row_number, row in enumerate(matrix[1:], start=2):
            if not any(str(value).strip() for value in row):
                continue
            source = {name: row[indexes[name]] for name in HOLDING_HEADERS}
            raw_rows.append(RawRow("Holdings", row_number, source, _row_hash(source)))
            values = {
                name: _decimal(source[name], name)
                for name in HOLDING_HEADERS if name != "Instrument"
            }
            if values["Qty."] <= 0:
                raise ValueError(f"non-positive holdings quantity at row {row_number}")
            variance = values["Invested"] - values["Qty."] * values["Avg. cost"]
            if abs(variance) > Decimal("0.01"):
                issues.append({
                    "severity": "INFO", "issue_type": "ROUNDED_AVERAGE_COST",
                    "row_number": row_number,
                    "evidence": {"variance": format(variance, "f")},
                })
            records.append(ParsedHolding(
                str(source["Instrument"]).strip().upper(), values["Qty."],
                values["Avg. cost"], values["LTP"], values["Invested"],
                values["Cur. val"], values["P&L"], values["Net chg."],
                values["Day chg."], row_number,
            ))
        totals = {
            key: format(sum((getattr(r, attr) for r in records), Decimal("0")), "f")
            for key, attr in (("invested", "invested"),
                              ("current_value", "current_value"), ("pnl", "pnl"))
        }
        return ParseResult(
            "holdings", self.format_version, tuple(raw_rows), tuple(records),
            {"rows": len(records), "totals": totals}, tuple(issues),
        )
