from __future__ import annotations

import os
import sqlite3
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
import duckdb
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from ai_trading_system.domains.trade_journal.analytics import behaviour_rate, score_components
from ai_trading_system.domains.trade_journal.enrichment import JournalMarketDataReader
from ai_trading_system.domains.trade_journal.importers import DhanHoldingsParser, DhanTradebookParser
from ai_trading_system.domains.trade_journal.service import TradeJournalService
from ai_trading_system.domains.trade_journal.store import TradeJournalStore
from ai_trading_system.ui.execution_api.app import create_app
from ai_trading_system.ui.execution_api.routes import trade_journal as journal_routes


HEADERS = [
    "Symbol", "ISIN", "Trade Date", "Exchange", "Segment", "Series", "Trade Type",
    "Auction", "Quantity", "Price", "Trade ID", "Order ID", "Order Execution Time",
]


def make_tradebook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equity"
    sheet.append(["Dhan account export"])
    sheet.append(["Generated for test"])
    sheet.append(HEADERS)
    sheet.append(["AAA", "INE012345678", datetime(2025, 1, 6), "NSE", "EQ", "EQ", "buy", False, 10, "100.25", "T1", "O1", datetime(2025, 1, 6, 10, 0)])
    sheet.append(["AAA", "INE012345678", datetime(2025, 1, 7), "NSE", "EQ", "EQ", "sell", False, 4, "110.00", "T2", "O2", datetime(2025, 1, 7, 10, 0)])
    workbook.save(path)
    return path


def test_header_discovery_import_replay_and_fifo(tmp_path: Path) -> None:
    path = make_tradebook(tmp_path / "tradebook.xlsx")
    parsed = DhanTradebookParser().parse(path)
    assert parsed.metadata["header_row"] == 3
    assert len(parsed.records) == 2

    store = TradeJournalStore(tmp_path, db_path=tmp_path / "journal.duckdb")
    store.migrate(apply=True)
    service = TradeJournalService(store)
    first = service.import_tradebook(path=path, broker="dhan", account_ref="account-a")
    second = service.import_tradebook(path=path, broker="dhan", account_ref="account-a")
    assert first.status == "IMPORTED"
    assert second.status == "NO_OP"
    journal_run_id = service.enqueue_task(action="reconstruct", account="account-a")
    assert service.run_task(journal_run_id)["status"] == "COMPLETED"
    with store.reader() as conn:
        assert conn.execute("SELECT count(*) FROM journal_fill").fetchone()[0] == 2
        assert conn.execute("SELECT count(*) FROM journal_order").fetchone()[0] == 2
        quantity, fifo_cost = conn.execute(
            "SELECT quantity,fifo_cost FROM journal_current_positions WHERE account_ref=?",
            ["account-a"],
        ).fetchone()
        assert conn.execute(
            "SELECT status FROM journal_task_request WHERE journal_run_id=?", [journal_run_id]
        ).fetchone()[0] == "COMPLETED"
    assert quantity == Decimal("6.00000000")
    assert fifo_cost == Decimal("601.50000000")


def test_holdings_parser_removes_blank_column_and_preserves_reported_totals(tmp_path: Path) -> None:
    path = tmp_path / "holdings.csv"
    path.write_text(
        "Instrument,Qty.,Avg. cost,LTP,Invested,Cur. val,P&L,Net chg.,Day chg.,\n"
        "AAA,2,10.00,12.00,20.02,24.00,3.98,19.88,1.00,\n",
        encoding="utf-8-sig",
    )
    parsed = DhanHoldingsParser().parse(path)
    assert len(parsed.records) == 1
    assert parsed.metadata["totals"] == {"invested": "20.02", "current_value": "24.00", "pnl": "3.98"}
    assert parsed.issues[0]["issue_type"] == "ROUNDED_AVERAGE_COST"


def test_holdings_identity_uses_operational_master_read_only(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    monkeypatch.setenv("DATA_ROOT", str(tmp_path))
    master = sqlite3.connect(tmp_path / "masterdata.db")
    master.execute(
        """CREATE TABLE symbols(symbol_id TEXT,nse_symbol TEXT,bse_symbol TEXT,
               isin TEXT,exchange TEXT)"""
    )
    master.execute(
        "INSERT INTO symbols VALUES (?,?,?,?,?)",
        ["BBB", "BBB", None, "INE012345679", "NSE"],
    )
    master.commit()
    master.close()
    holdings = tmp_path / "holdings.csv"
    holdings.write_text(
        "Instrument,Qty.,Avg. cost,LTP,Invested,Cur. val,P&L,Net chg.,Day chg.\n"
        "BBB,2,10,12,20,24,4,20,1\n",
        encoding="utf-8-sig",
    )
    store = TradeJournalStore(Path.cwd(), db_path=tmp_path / "journal.duckdb")
    store.migrate(apply=True)
    TradeJournalService(store).import_holdings(
        path=holdings, broker="dhan", account_ref="account-a",
        as_of=date(2025, 1, 8), mode="reconciliation_only",
    )
    with store.reader() as conn:
        resolved = conn.execute(
            "SELECT instrument_id FROM portfolio_snapshot_position WHERE instrument='BBB'"
        ).fetchone()[0]
        method = conn.execute(
            "SELECT method FROM identity_resolution WHERE instrument_id=?", [resolved]
        ).fetchone()[0]
    assert resolved.startswith("ins_")
    assert method == "operational_master"


def test_holdings_modes_do_not_confuse_checkpoint_with_broker_fill(tmp_path: Path) -> None:
    tradebook = make_tradebook(tmp_path / "tradebook.xlsx")
    holdings = tmp_path / "holdings.csv"
    holdings.write_text(
        "Instrument,Qty.,Avg. cost,LTP,Invested,Cur. val,P&L,Net chg.,Day chg.,\n"
        "AAA,6,100.25,110,601.50,660,58.50,9.73,0,\n",
        encoding="utf-8-sig",
    )
    (tmp_path / "data").mkdir()
    master = sqlite3.connect(tmp_path / "data" / "masterdata.db")
    master.execute(
        """CREATE TABLE symbols(symbol_id TEXT,nse_symbol TEXT,bse_symbol TEXT,
               isin TEXT,exchange TEXT)"""
    )
    master.execute(
        "INSERT INTO symbols VALUES (?,?,?,?,?)",
        ["AAA", "AAA", None, "INE012345678", "NSE"],
    )
    master.commit()
    master.close()
    store = TradeJournalStore(tmp_path, db_path=tmp_path / "journal.duckdb")
    store.migrate(apply=True)
    service = TradeJournalService(store)
    service.import_tradebook(path=tradebook, broker="dhan", account_ref="account-a")
    service.import_holdings(
        path=holdings, broker="dhan", account_ref="account-a",
        as_of=date(2025, 1, 8), mode="reconciliation_only",
    )
    repeated = service.reconcile_latest("account-a")
    assert repeated["status"] == "NO_OP"
    with store.reader() as conn:
        assert conn.execute(
            "SELECT count(*) FROM portfolio_event WHERE account_ref=? AND event_type='OPENING'",
            ["account-a"],
        ).fetchone()[0] == 0

    service.import_holdings(
        path=holdings, broker="dhan", account_ref="account-b",
        as_of=date(2025, 1, 5), mode="opening_anchor",
    )
    service.reconstruct("account-b")
    with store.reader() as conn:
        opening_count = conn.execute(
            "SELECT count(*) FROM opening_position WHERE account_ref=? AND provenance='snapshot_bootstrap'",
            ["account-b"],
        ).fetchone()[0]
        position = conn.execute(
            "SELECT quantity,fifo_cost FROM journal_current_positions WHERE account_ref=?",
            ["account-b"],
        ).fetchone()
    assert opening_count == 1
    assert position == (Decimal("6.00000000"), Decimal("601.50000000"))


def test_analytics_coverage_and_minimum_sample_gates() -> None:
    assert score_components({"a": Decimal("80"), "b": None})["status"] == "insufficient_data"
    assert behaviour_rate(eligible=4, occurrences=3)["status"] == "insufficient_sample"
    finding = behaviour_rate(eligible=10, occurrences=4)
    assert finding["status"] == "reportable"
    assert finding["occurrences"] == 4 and finding["eligible"] == 10


def test_market_context_is_strictly_previous_session(tmp_path: Path) -> None:
    db_path = tmp_path / "ohlcv.duckdb"
    conn = duckdb.connect(str(db_path))
    conn.execute(
        """CREATE TABLE _catalog(symbol_id VARCHAR,exchange VARCHAR,timestamp TIMESTAMP,
               open DOUBLE,high DOUBLE,low DOUBLE,close DOUBLE,volume DOUBLE,
               provider VARCHAR,validation_status VARCHAR,ingestion_version VARCHAR)"""
    )
    conn.execute(
        """INSERT INTO _catalog VALUES
           ('AAA','NSE','2025-01-05',99,102,98,100,1000,'nse','trusted','v1'),
           ('AAA','NSE','2025-01-06',100,1000,1,999,9999,'nse','trusted','v1')"""
    )
    conn.close()
    context = JournalMarketDataReader(db_path).prior_session_context(
        symbol="AAA", exchange="NSE", decision_date=date(2025, 1, 6)
    )
    assert context["cutoff_session"] == "2025-01-05"
    assert context["metrics"]["close"] == 100


def test_versioned_analysis_persists_outcomes_and_valuation_series(tmp_path: Path) -> None:
    market_path = tmp_path / "ohlcv.duckdb"
    conn = duckdb.connect(str(market_path))
    conn.execute(
        """CREATE TABLE _catalog(symbol_id VARCHAR,exchange VARCHAR,timestamp TIMESTAMP,
               open DOUBLE,high DOUBLE,low DOUBLE,close DOUBLE,volume DOUBLE,
               provider VARCHAR,validation_status VARCHAR,ingestion_version VARCHAR)"""
    )
    start = date(2024, 1, 1)
    rows = []
    for index in range(430):
        session = start + timedelta(days=index)
        close = 80 + index * 0.1
        rows.append(("AAA", "NSE", session, close - 1, close + 2, close - 2, close, 1000 + index, "nse", "trusted_primary", "v1"))
    conn.executemany("INSERT INTO _catalog VALUES (?,?,?,?,?,?,?,?,?,?,?)", rows)
    conn.execute(
        """CREATE TABLE _corporate_actions(isin VARCHAR,symbol VARCHAR,ex_date DATE,
               action_type VARCHAR,share_factor DOUBLE,price_factor DOUBLE,source VARCHAR,
               action_key VARCHAR,raw_payload_hash VARCHAR,parsed_ratio VARCHAR,status VARCHAR)"""
    )
    conn.execute(
        """INSERT INTO _corporate_actions VALUES
           ('INE012345678','AAA','2025-01-08','split',2,0.5,'nse_corporate_actions',
            'action-1','source-hash','1:2','active')"""
    )
    conn.close()
    store = TradeJournalStore(tmp_path, db_path=tmp_path / "journal.duckdb")
    store.migrate(apply=True)
    service = TradeJournalService(store)
    service.import_tradebook(
        path=make_tradebook(tmp_path / "tradebook.xlsx"), broker="dhan", account_ref="account-a"
    )
    result = service.analyze("account-a", market_data=JournalMarketDataReader(market_path))
    assert result["status"] == "COMPLETED"
    assert result["evaluations"] == 2
    assert result["valuation_sessions"] > 0
    assert result["corporate_action_proposals"] == 1
    with store.reader() as conn:
        payload = conn.execute(
            "SELECT components_json FROM trade_evaluation WHERE evaluation_type='ENTRY_PROCESS'"
        ).fetchone()[0]
        assert '"starts_next_session":true' in payload
        assert conn.execute("SELECT count(*) FROM portfolio_valuation").fetchone()[0] > 0
        assert conn.execute("SELECT scope_label FROM portfolio_evaluation").fetchone()[0] == "holdings_only"
        assert conn.execute(
            "SELECT review_status FROM corporate_action_event"
        ).fetchone()[0] == "PROPOSED"
        assert conn.execute(
            "SELECT quantity FROM journal_current_positions"
        ).fetchone()[0] == Decimal("6.00000000")


def test_deficit_fill_never_becomes_a_short_valuation(tmp_path: Path) -> None:
    tradebook = tmp_path / "deficit.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equity"
    sheet.append(HEADERS)
    sheet.append(["AAA", "INE012345678", datetime(2025, 1, 6), "NSE", "EQ", "EQ", "sell", False, 10, "100", "T0", "O0", datetime(2025, 1, 6, 10)])
    sheet.append(["AAA", "INE012345678", datetime(2025, 1, 7), "NSE", "EQ", "EQ", "buy", False, 5, "101", "T1", "O1", datetime(2025, 1, 7, 10)])
    workbook.save(tradebook)
    market_path = tmp_path / "ohlcv.duckdb"
    conn = duckdb.connect(str(market_path))
    conn.execute(
        """CREATE TABLE _catalog(symbol_id VARCHAR,exchange VARCHAR,timestamp TIMESTAMP,
               open DOUBLE,high DOUBLE,low DOUBLE,close DOUBLE,volume DOUBLE,
               provider VARCHAR,validation_status VARCHAR,ingestion_version VARCHAR)"""
    )
    conn.executemany(
        "INSERT INTO _catalog VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        [("AAA", "NSE", date(2025, 1, day), 99, 102, 98, 100 + day, 1000, "nse", "trusted_primary", "v1") for day in range(1, 15)],
    )
    conn.close()
    store = TradeJournalStore(tmp_path, db_path=tmp_path / "journal.duckdb")
    store.migrate(apply=True)
    service = TradeJournalService(store)
    imported = service.import_tradebook(path=tradebook, broker="dhan", account_ref="account-a")
    assert imported.summary["deficits"] == 1
    service.analyze("account-a", market_data=JournalMarketDataReader(market_path))
    with store.reader() as conn:
        assert conn.execute("SELECT min(quantity) FROM portfolio_valuation").fetchone()[0] >= 0
        metrics = conn.execute("SELECT metrics_json FROM portfolio_evaluation").fetchone()[0]
    assert '"top_1_weight":"1"' in metrics
    assert '"trust_status":"PARTIAL"' in metrics


def test_reviewed_corporate_action_transforms_quantity_not_total_cost(tmp_path: Path) -> None:
    store = TradeJournalStore(tmp_path, db_path=tmp_path / "journal.duckdb")
    store.migrate(apply=True)
    service = TradeJournalService(store)
    service.import_tradebook(
        path=make_tradebook(tmp_path / "tradebook.xlsx"),
        broker="dhan", account_ref="account-a",
    )
    with store.reader() as conn:
        instrument_id = conn.execute(
            "SELECT instrument_id FROM journal_fill LIMIT 1"
        ).fetchone()[0]
    proposal = service.propose_corporate_action(
        instrument_id=instrument_id, action_type="split", effective_date=date(2025, 1, 8),
        quantity_factor=Decimal("2"), cost_factor=Decimal("0.5"), source_ref="reviewed-test",
    )
    service.approve_corporate_action(proposal["action_id"], reviewer="operator")
    with store.reader() as conn:
        quantity, fifo_cost = conn.execute(
            "SELECT quantity,fifo_cost FROM journal_current_positions WHERE account_ref=?",
            ["account-a"],
        ).fetchone()
    assert quantity == Decimal("12.00000000")
    assert fifo_cost == Decimal("601.50000000")


def test_execution_api_auth_preview_and_decimal_strings(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setenv("DATA_ROOT", str(tmp_path))
    monkeypatch.setenv("AI_TRADING_PROJECT_ROOT", str(Path.cwd()))
    monkeypatch.setenv("EXECUTION_API_KEY", "journal-test-key")
    TradeJournalStore(Path.cwd()).migrate(apply=True)
    path = make_tradebook(tmp_path / "api-tradebook.xlsx")
    client = TestClient(create_app())
    assert client.get("/api/trade-journal/accounts").status_code == 401
    with path.open("rb") as handle:
        preview = client.post(
            "/api/trade-journal/imports/tradebook/preview",
            headers={"x-api-key": "journal-test-key"},
            files={"file": (path.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert preview.status_code == 200
    sha = preview.json()["file_sha256"]
    with path.open("rb") as handle:
        committed = client.post(
            "/api/trade-journal/imports/tradebook/commit",
            headers={"x-api-key": "journal-test-key"},
            data={"broker": "dhan", "account_ref": "account-a", "expected_sha256": sha},
            files={"file": (path.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert committed.status_code == 201
    replay_path = make_tradebook(tmp_path / "api-tradebook-replay.xlsx")
    replay_workbook = load_workbook(replay_path)
    replay_workbook["Equity"]["A1"] = "Dhan account export replay"
    replay_workbook.save(replay_path)
    replay_result = TradeJournalService(TradeJournalStore(Path.cwd())).import_tradebook(
        path=replay_path, broker="dhan", account_ref="account-a"
    )
    assert replay_result.status == "IMPORTED"
    first_page = client.get(
        "/api/trade-journal/imports?account_ref=account-a&limit=1",
        headers={"x-api-key": "journal-test-key"},
    ).json()
    assert first_page["pagination"]["has_more"] is True
    second_page = client.get(
        f"/api/trade-journal/imports?account_ref=account-a&limit=1&cursor={first_page['pagination']['next_cursor']}",
        headers={"x-api-key": "journal-test-key"},
    ).json()
    assert second_page["items"][0]["import_id"] != first_page["items"][0]["import_id"]
    assert client.get(
        "/api/trade-journal/imports?cursor=not-a-valid-cursor",
        headers={"x-api-key": "journal-test-key"},
    ).status_code == 422
    positions = client.get(
        "/api/trade-journal/positions?account_ref=account-a",
        headers={"x-api-key": "journal-test-key"},
    )
    assert positions.status_code == 200
    assert positions.json()["items"][0]["quantity"] == "6.00000000"
    episodes = client.get(
        "/api/trade-journal/episodes?account_ref=account-a",
        headers={"x-api-key": "journal-test-key"},
    ).json()["items"]
    annotation = client.post(
        "/api/trade-journal/annotations",
        headers={"x-api-key": "journal-test-key"},
        data={"episode_id": episodes[0]["episode_id"], "thesis": "Test thesis", "tags_json": '["test"]'},
    )
    assert annotation.status_code == 201
    assert annotation.json()["revision"] == 1
    opening = client.post(
        "/api/trade-journal/opening-lots/propose",
        headers={"x-api-key": "journal-test-key"},
        data={
            "account_ref": "account-a", "instrument_id": positions.json()["items"][0]["instrument_id"],
            "effective_at": "2025-01-05T00:00:00", "quantity": "1", "total_cost": "90",
            "reason": "reviewed pre-period inventory",
        },
    )
    assert opening.status_code == 201
    adjustment_id = opening.json()["adjustment_id"]
    approved = client.post(
        f"/api/trade-journal/opening-lots/{adjustment_id}/approve",
        headers={"x-api-key": "journal-test-key"}, data={"reviewer": "operator-test"},
    )
    assert approved.status_code == 200
    assert client.post(
        f"/api/trade-journal/opening-lots/{adjustment_id}/approve",
        headers={"x-api-key": "journal-test-key"}, data={"reviewer": "operator-test"},
    ).status_code == 409
    launched: dict[str, object] = {}

    def fake_launch(**kwargs: object) -> str:
        launched.update(kwargs)
        return "task-journal-test"

    monkeypatch.setattr(journal_routes, "_launch_subprocess_task", fake_launch)
    task = client.post(
        "/api/trade-journal/reconstructions",
        headers={"x-api-key": "journal-test-key"},
        data={"account_ref": "account-a"},
    )
    assert task.status_code == 202
    assert task.json()["task_id"] == "task-journal-test"
    command = [str(value) for value in launched["command"]]  # type: ignore[index]
    assert "account-a" not in command
    assert "--journal-run-id" in command


@pytest.mark.skipif(
    not os.getenv("TRADE_JOURNAL_SAMPLE_TRADEBOOK") or not os.getenv("TRADE_JOURNAL_SAMPLE_HOLDINGS"),
    reason="operator sample paths are opt-in",
)
def test_operator_file_characterization(tmp_path: Path) -> None:
    tradebook = Path(os.environ["TRADE_JOURNAL_SAMPLE_TRADEBOOK"])
    holdings = Path(os.environ["TRADE_JOURNAL_SAMPLE_HOLDINGS"])
    trade_parsed = DhanTradebookParser().parse(tradebook)
    hold_parsed = DhanHoldingsParser().parse(holdings)
    assert len(trade_parsed.records) == 5_261
    assert sum(row.side == "buy" for row in trade_parsed.records) == 2_927
    assert sum(row.side == "sell" for row in trade_parsed.records) == 2_334
    assert len(trade_parsed.issues) == 15
    assert len(hold_parsed.records) == 14
    assert {key: Decimal(value) for key, value in hold_parsed.metadata["totals"].items()} == {
        "invested": Decimal("7892527.99"),
        "current_value": Decimal("8717607.50"),
        "pnl": Decimal("825079.51"),
    }
    store = TradeJournalStore(tmp_path, db_path=tmp_path / "characterization.duckdb")
    store.migrate(apply=True)
    result = TradeJournalService(store).import_tradebook(
        path=tradebook, broker="dhan", account_ref="characterization-account"
    )
    assert result.summary["deficits"] == 19
    with store.reader() as conn:
        assert conn.execute("SELECT count(*) FROM journal_order").fetchone()[0] == 827
        assert conn.execute("SELECT max(fill_count) FROM journal_order").fetchone()[0] == 79
